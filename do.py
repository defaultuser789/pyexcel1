# 脚本实现对excel 表格的处理，处理完成后另存为新文件

import openpyxl

# 打开excel 表格
workbook = openpyxl.load_workbook("公司级系统清单-1.xlsx")

# 指定sheet 表
sheet = workbook["访问公司系统统计"]

# 如果不想要在一个任务中跑所有的扫描，可以修改此处的数值，即为开始和结束的行号

# column_number 定义要写入的列号，这里使用数字，速查表：

# 1  2  3  4  5  6  7  8  9  10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 26
# A  B  C  D  E  F  G  H  I  J  K  L  M  N  O  P  Q  R  S  T  U  V  W  X  Y  Z

start_column_number = 9
stop_column_number = sheet.max_column

start_row_number = 2
stop_row_number = sheet.max_row

letter_tuple = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z')

for column_number in range(start_column_number, stop_column_number + 1, 1):

    for row_number in range(start_row_number, stop_row_number + 1, 1):
    
        cell = sheet.cell(row=row_number, column=column_number)
        cell_value = str(cell.value).split("\n")
        result = ""
        
        # 打印处理进度信息
        print('正在处理第',letter_tuple[column_number - 1],'列','第',row_number,'行')
        
        for value in cell_value:
            # 单元格内容以行为单位进行处理，功能是将包含filtered 和closed 的行删除，并替换某些字符串
            if value.find("filtered") == -1 and value.find("closed") == -1:
                result = result + value.replace("protocol: ", "").replace("  port: ", "/").replace("  state: ", " : ") + "\n"

        # 最终结果写入excel 表中
        sheet.cell(row=row_number, column=column_number).value = result

        # 另存为新文件
        workbook.save("公司级系统清单-2.xlsx")